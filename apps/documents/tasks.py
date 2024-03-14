import logging
from collections import OrderedDict
from datetime import datetime
from decimal import Decimal
from io import BytesIO

import carbone_sdk
from dateutil.relativedelta import relativedelta
from django.conf import settings
from django.core.files.base import ContentFile
from django.db.models import DecimalField, Max, Min, Q, Sum
from django.db.models.functions import Coalesce
from django.utils import timezone
from invoices.models import BillingDetail, IndustryCharges, MeterInvoice
from invoices.utils import (
    INVOICE_EXCEL_COLUMNS,
    compute_industry_charges_total_charges,
    exclude_electricity_charges_fields,
    get_electricity_charges,
    write_excel_row,
)
from openpyxl import load_workbook
from sources.billing_service.models import BillingServiceContract
from sources.customer_portal.enums import ElectricityBillStatusEnum
from sources.customer_portal.models import (
    CustomerPortalECA,
    CustomerPortalElectricityBillMeter,
    CustomerPortalElectricityContract,
)
from sources.data_service.models import DataServiceMPAN
from templates.enums import TemplateFileTypesEnum, TemplateSubTypesEnum
from templates.models import TemplateFile

from celery import shared_task

from .models import DataFileRequest, Document, DocumentGenerationRequest
from .serializers import DataFileRequestDetailSerializer
from .utils import map_data

logger = logging.getLogger(__name__)


@shared_task
def process_document_generation_request(document_id):
    document = Document.objects.get(id=document_id)
    template_data_mapping = document.template_data_mapping
    mapping_expression = template_data_mapping.mapping_expression
    template_id = template_data_mapping.template_file.external_id
    csdk_test = carbone_sdk.CarboneSDK(settings.CARBONE_IO_API_TEST_TOKEN)
    csdk_production = carbone_sdk.CarboneSDK(
        settings.CARBONE_IO_API_PRODUCTION_TOKEN
    )

    for entry in document.generated_documents.all():
        data = map_data(entry.json_data, mapping_expression)
        json_data = {
            "data": data,
            "convertTo": "pdf",
            "lang": "en-gb",
        }

        try:
            if entry.is_production:
                report_bytes, _ = csdk_production.render(template_id, json_data)
            else:
                report_bytes, _ = csdk_test.render(template_id, json_data)
        except Exception as err:
            entry.status = DocumentGenerationRequest.FAILED
            entry.error = err
            entry.save()

            continue

        try:
            entry.file.save(
                f"{document.name} ({entry.name}).pdf",
                ContentFile(report_bytes),
                save=True,
            )
        except Exception:
            continue

        entry.status = DocumentGenerationRequest.SUCCEEDED
        entry.save()


def create_data_file_excel(
    data_file_request: DataFileRequest,
    excel_template_file: TemplateFile,
):
    wb = load_workbook(excel_template_file.file.path)
    ws = wb.active
    row_number = 1

    headers = exclude_electricity_charges_fields(
        INVOICE_EXCEL_COLUMNS,
        [cell.value for cell in ws[1]],
    )

    row_number += 1

    data = DataFileRequestDetailSerializer(data_file_request).data
    columns = list(headers.values())

    for index, meter_invoice in enumerate(data["meter_invoices"]):
        write_excel_row(ws, row_number, columns, meter_invoice)
        for related_field in [
            "customer_billing_details",
            "hh_consumption_charges",
            "reading_consumption_charges",
        ]:
            if meter_invoice.get(related_field):
                write_excel_row(
                    ws,
                    row_number,
                    columns,
                    meter_invoice[related_field],
                )

        if meter_invoice.get("industry_charges"):
            data = OrderedDict()
            total_industry_charges = 0
            for index, industry_charge in enumerate(
                meter_invoice["industry_charges"]
            ):
                for field in IndustryCharges._meta.fields:
                    field_value = ""
                    if industry_charge.get(field.name) is not None:
                        field_value = industry_charge[field.name]

                        if field.name == "charges" and industry_charge.get(
                            field.name
                        ):
                            total_industry_charges += Decimal(
                                industry_charge[field.name]
                            )
                    data[
                        f"industry_charge_{index + 1}_{field.name}"
                    ] = field_value
            data["total_industry_charges"] = total_industry_charges
            write_excel_row(
                ws,
                row_number,
                columns,
                data,
            )

        row_number += 1

    excel_data = BytesIO()
    wb.save(excel_data)
    excel_data.seek(0)

    data_file_request.xls_file.save(
        f"{data_file_request.name} {timezone.now():%Y%m%d_%H%M%S}.xlsx",
        ContentFile(excel_data.read()),
    )


@shared_task
def process_data_file(data_file_id: int) -> None:
    data_file_request = DataFileRequest.objects.get(pk=data_file_id)
    template_sub_type = data_file_request.document_template.sub_type
    period_start_at = data_file_request.period_start_at
    period_end_at = data_file_request.period_end_at
    today = timezone.now()

    filters = {}
    contract_filters = {}
    meter_filters = {}
    if data_file_request.electricity_customer_accounts:
        filters["pk__in"] = [
            data["id"]
            for data in data_file_request.electricity_customer_accounts
        ]
    if data_file_request.account_holders:
        filters["account_holder_id__in"] = [
            data["id"] for data in data_file_request.account_holders
        ]
    if data_file_request.customers:
        filters["account_holder__customer_id__in"] = [
            data["id"] for data in data_file_request.customers
        ]
    if data_file_request.contracts:
        contract_filters["pk__in"] = [
            data["id"] for data in data_file_request.contracts
        ]
    if data_file_request.sites:
        meter_filters["asset_id__in"] = [
            data["id"] for data in data_file_request.sites
        ]
    if data_file_request.mpans:
        meter_filters["mpan_id__in"] = [
            data["id"] for data in data_file_request.mpans
        ]

    for electricity_customer_account in CustomerPortalECA.objects.filter(
        **filters
    ):
        customer_portal_customer_id = (
            electricity_customer_account.account_holder.customer_id
        )
        meters_consumption = electricity_customer_account.active_meters.filter(
            is_smart_meter=template_sub_type.name == TemplateSubTypesEnum.HH,
            **meter_filters,
        ).annotate(
            consumption=Sum(
                "consumptions__consumption",
                filter=Q(
                    consumptions__created_at__date__gte=period_start_at,
                    consumptions__created_at__date__lte=period_end_at,
                ),
            ),
            cost=Sum(
                "consumptions__cost",
                filter=Q(
                    consumptions__created_at__date__gte=period_start_at,
                    consumptions__created_at__date__lte=period_end_at,
                ),
            ),
            opening_reading=Min(
                "consumptions__reading",
                filter=Q(
                    consumptions__created_at__date__gte=period_start_at,
                    consumptions__created_at__date__lte=period_end_at,
                ),
            ),
            opening_reading_at=Min(
                "consumptions__created_at",
                filter=Q(
                    consumptions__created_at__date__gte=period_start_at,
                    consumptions__created_at__date__lte=period_end_at,
                ),
            ),
            last_reading=Max(
                "consumptions__reading",
                filter=Q(
                    consumptions__created_at__date__gte=period_start_at,
                    consumptions__created_at__date__lte=period_end_at,
                ),
            ),
            last_reading_at=Max(
                "consumptions__created_at",
                filter=Q(
                    consumptions__created_at__date__gte=period_start_at,
                    consumptions__created_at__date__lte=period_end_at,
                ),
            ),
        )

        for index, meter in enumerate(meters_consumption):
            # TODO: splitting of HH and non-HH meters
            active_contracts = CustomerPortalElectricityContract.objects.filter(
                is_active=True,
                end_date__gt=today,
                account__active_meters__pk=meter.id,
                **contract_filters,
            ).distinct()
            for contract in active_contracts:
                try:
                    billing_service_contract = BillingServiceContract.objects.get(
                        account__customer__customer_portal_id=customer_portal_customer_id,
                        name=contract.name,
                        customer_portal_id=contract.id,
                    )
                except (
                    BillingServiceContract.DoesNotExist,
                    BillingServiceContract.MultipleObjectsReturned,
                ) as e:
                    logger.error(
                        f"Contract exception for MPAN: {meter.mpan.mpan} -- {e}"
                    )
                    continue

                contract_name = (
                    billing_service_contract.name.replace("-", "")
                    if billing_service_contract.name
                    else f"{billing_service_contract.id}"
                )
                invoice_number = f"{contract_name}{today:%y%m}-{index + 1}"

                meter_mpan = getattr(meter.mpan, "mpan", "")
                data_service_mpan = DataServiceMPAN.objects.filter(
                    mpan=meter_mpan
                )
                data_service_mpan = (
                    data_service_mpan.values("mtc", "pc", "llfc").first()
                    if data_service_mpan.exists()
                    else {}
                )

                customer_billing_details = BillingDetail.objects.create(
                    invoice_number=invoice_number,
                    billing_name=billing_service_contract.account.billing_name,
                    billing_address_1=billing_service_contract.account.billing_address_1,
                    billing_address_2=billing_service_contract.account.billing_address_2,
                    billing_address_3=billing_service_contract.account.billing_address_3,
                    billing_address_4=billing_service_contract.account.billing_address_4,
                    billing_address_5=billing_service_contract.account.billing_address_5,
                    billing_city=billing_service_contract.account.billing_city,
                    billing_postal_code=billing_service_contract.account.billing_postal_code,
                    site_name=getattr(meter.asset, "name", ""),
                    site_address=getattr(meter.asset, "address", ""),
                    vat_number=getattr(
                        billing_service_contract, "vat_number", ""
                    ),
                    account_number=getattr(contract, "name", "") or "",
                    msn=getattr(meter.device, "serial_number", ""),
                    mpan=meter_mpan,
                    mtc=data_service_mpan.get("mtc"),
                    pc=data_service_mpan.get("pc"),
                    llf=data_service_mpan.get("llfc"),
                    invoice_at=today,
                    bill_from_at=(
                        timezone.make_aware(
                            datetime.combine(
                                period_start_at, datetime.min.time()
                            )
                        )
                    ),
                    bill_to_at=(
                        timezone.make_aware(
                            datetime.combine(period_end_at, datetime.min.time())
                        )
                    ),
                    contract_end_at=contract.end_date,
                    payment_due_at=period_start_at + relativedelta(days=15),
                )

                previous_amount = (
                    CustomerPortalElectricityBillMeter.objects.filter(
                        meter=meter,
                        bill__status__in=[
                            ElectricityBillStatusEnum.PENDING,
                            ElectricityBillStatusEnum.OVERDUE,
                        ],
                    ).aggregate(
                        previous_amount=Coalesce(
                            Sum("total"), 0, output_field=DecimalField()
                        ),
                    )[
                        "previous_amount"
                    ]
                )

                total_no_vat = meter.cost or 0
                charged_vat = total_no_vat * (
                    billing_service_contract.vat / 100
                )
                bill_amount = total_no_vat + charged_vat

                meter_invoice = MeterInvoice.objects.create(
                    data_file_request=data_file_request,
                    customer_billing_details=customer_billing_details,
                    total_no_vat=total_no_vat,
                    applicable_vat=billing_service_contract.vat / 100,
                    charged_vat=charged_vat,
                    previous_balance=previous_amount,
                    bill_amount=bill_amount,
                    total_amount=previous_amount + bill_amount,
                )

                get_electricity_charges(
                    meter,
                    meter_invoice,
                    billing_service_contract,
                    customer_portal_customer_id,
                )

    compute_industry_charges_total_charges(data_file_request)
    excel_template_file = (
        data_file_request.document_template.template_files.filter(
            file_type=TemplateFileTypesEnum.XLS, is_active=True
        ).last()
    )
    if excel_template_file and excel_template_file.file:
        create_data_file_excel(data_file_request, excel_template_file)
